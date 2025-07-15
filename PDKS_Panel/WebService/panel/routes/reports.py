from flask import Blueprint, render_template, request, jsonify, send_file, make_response, flash, redirect, url_for
from flask_login import login_required, current_user
from functools import wraps
from sqlalchemy import func, and_, or_, cast, String, extract, text, case, DateTime
from sqlalchemy.orm import joinedload
from panel.models import User, Device, Record, Department
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker, scoped_session
import os

# Database bağlantısı
db_user = os.environ.get('DB_USER', 'dbuser')
db_password = os.environ.get('DB_PASSWORD', 'dbpass123')
db_host = os.environ.get('DB_HOST', 'pdks_database')
db_port = os.environ.get('DB_PORT', '5433')
db_name = os.environ.get('DB_NAME', 'myapp_db')

DATABASE_URL = f"postgresql://{db_user}:{db_password}@{db_host}:{db_port}/{db_name}"
engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(bind=engine)
db_session = scoped_session(SessionLocal)

from datetime import datetime, timedelta, date
from decimal import Decimal
import io
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.fonts import addMapping
from reportlab.pdfbase.pdfmetrics import registerFont
from reportlab.lib.units import inch
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json
import re
import logging

# Logger konfigürasyonu
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

reports = Blueprint('reports', __name__, url_prefix='/reports')

def admin_required(f):
    """Sadece Admin (role_id = 1) kullanıcıları için decorator"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        
        if current_user.role_id != 1:
            logger.warning(f"Admin yetkisi gerekli - User: {current_user.email}, Role: {current_user.role_id}")
            flash('Yetkisiz Erişim! Bu sayfaya erişim yetkiniz bulunmamaktadır.', 'danger')
            return redirect(url_for('index'))
        
        return f(*args, **kwargs)
    return decorated_function

# Sütun başlıkları ve sıralama için sabit tanımlamalar
COLUMN_HEADERS = {
    'user_id': 'Sicil No',
    'name': 'Ad Soyad',
    'card': 'Kart No',
    'department': 'Departman',
    'occupation': 'Meslek',
    'job_title': 'İş Tanımı',
    'work_type': 'Çalışma Tipi',
    'employment_status': 'Çalışma Durumu',
    'tarih': 'Tarih',
    'saat': 'Saat',
    'islem': 'İşlem',
    'kayit_tipi': 'Terminal',
    'personel_adi': 'Ad Soyad',
    'ilk_giris': 'İlk Giriş',
    'son_cikis': 'Son Çıkış',
    'toplam_sure': 'Toplam Süre',
    'kayit_sayisi': 'Kayıt Sayısı',
    'hafta_baslangic': 'Hafta Başlangıç',
    'calisilan_gun': 'Çalışılan Gün',
    'week_number': 'Hafta No',
    'Açıklama': 'Açıklama'
}

COLUMN_ORDER = [
    'user_id',
    'name',
    'card',
    'department',
    'occupation',
    'job_title',
    'work_type',
    'employment_status'
]

REPORT_TITLES = {
    'personnel_list': 'Personel_Listesi',
    'detailed_attendance': 'Detayli_Giris_Cikis',
    'detailed_timesheet': 'Detayli_Puantaj',
    'weekly_timesheet': 'Haftalik_Puantaj',
    'daily_timesheet': 'Gunluk_Puantaj'
}

@reports.route('/')
@login_required
@admin_required
def reports_page():
    try:
        # Departmanları al
        departments = db_session.query(Department.department_name)\
            .filter(Department.deleted_at.is_(None))\
            .order_by(Department.department_name)\
            .all()
        departments = [dept[0] for dept in departments]

        # Personel listesini al
        employees = db_session.query(User.user_id, User.name)\
            .filter(User.deleted_at.is_(None))\
            .order_by(User.name)\
            .all()

        return render_template('reports/reports.html',
                             departments=departments,
                             employees=employees)
    except Exception as e:
        logger.error(f"Raporlar sayfası hatası: {str(e)}")
        flash('Bir hata oluştu!', 'danger')
        return render_template('reports/reports.html')
    finally:
        db_session.remove()

def convert_timedelta(obj):
    """Timedelta ve diğer özel tipleri string'e dönüştürür"""
    if obj is None:
        return '-'
    elif isinstance(obj, timedelta):
        total_seconds = int(obj.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return f"{hours:02d}:{minutes:02d}"
    elif isinstance(obj, Decimal):
        return str(obj)
    elif isinstance(obj, datetime):
        return obj.strftime('%d.%m.%Y %H:%M')
    elif isinstance(obj, date):
        return obj.strftime('%d.%m.%Y')
    elif obj == "":
        return '-'
    elif str(obj).lower() in ('none', 'null', 'nan'):
        return '-'
    return obj

def strip_html_tags(text):
    """HTML etiketlerini metinden temizler"""
    if not isinstance(text, str):
        return text
    # HTML etiketlerini temizle
    clean_text = re.sub(r'<[^>]*>', '', text)
    # HTML entity'leri de düzelt (&lt; gibi)
    clean_text = clean_text.replace('&lt;', '<').replace('&gt;', '>').replace('&amp;', '&')
    return clean_text

def convert_data_for_json(data):
    """Veri yapısındaki tüm timedelta değerlerini dönüştürür"""
    if isinstance(data, dict):
        return {key: convert_data_for_json(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [convert_data_for_json(item) for item in data]
    elif hasattr(data, '_asdict'):  # psycopg2.extras.DictRow nesneleri için
        return {key: convert_data_for_json(value) for key, value in data._asdict().items()}
    else:
        return convert_timedelta(data)

@reports.route('/generate', methods=['POST'])
@login_required
@admin_required
def generate():
    try:
        # Form verilerini al
        report_type = request.form.get('report_type')
        user_id = request.form.get('user_id')
        department = request.form.get('departman')  # HTML formunda 'departman' olarak tanımlı
        occupation = request.form.get('meslek')     # HTML formunda 'meslek' olarak tanımlı
        egitim_durumu = request.form.get('egitim_durumu')
        work_type = request.form.get('calisma_tipi')  # HTML formunda 'calisma_tipi' olarak tanımlı
        employment_status = request.form.get('calisma_durumu')  # HTML formunda 'calisma_durumu' olarak tanımlı
        
        # Tarih verilerini al - hem date_range hem de ayrı tarihler desteklensin
        date_range = request.form.get('date_range')
        form_start_date = request.form.get('start_date')
        form_end_date = request.form.get('end_date')

        # Tarih aralığını parse et
        start_date = end_date = None
        if report_type != 'personnel_list':
            try:
                if date_range:
                    # Eski format: date_range varsa onu kullan
                    start_date_str, end_date_str = date_range.split(' - ')
                    start_date = datetime.strptime(start_date_str, '%d/%m/%Y').strftime('%Y-%m-%d')
                    end_date = datetime.strptime(end_date_str, '%d/%m/%Y').strftime('%Y-%m-%d')
                elif form_start_date and form_end_date:
                    # Yeni format: ayrı tarih alanları varsa onları kullan
                    # HTML date input'u zaten YYYY-MM-DD formatında gelir
                    start_date = form_start_date
                    end_date = form_end_date
                    
                    # Tarih formatını kontrol et ve gerekirse düzelt
                    try:
                        # Gelen tarihlerin formatını kontrol et
                        datetime.strptime(start_date, '%Y-%m-%d')
                        datetime.strptime(end_date, '%Y-%m-%d')
                    except ValueError:
                        # Farklı format gelirse dönüştürmeyi dene
                        try:
                            start_date = datetime.strptime(start_date, '%d/%m/%Y').strftime('%Y-%m-%d')
                            end_date = datetime.strptime(end_date, '%d/%m/%Y').strftime('%Y-%m-%d')
                        except ValueError:
                            # Hiçbir format çalışmazsa varsayılan tarihleri kullan
                            logger.warning("Tarih formatı tanınamadı, varsayılan tarihler kullanılıyor")
                            end_date = datetime.now().strftime('%Y-%m-%d')
                            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
                else:
                    # Hiçbir tarih verilmemişse varsayılan aralığı kullan
                    end_date = datetime.now().strftime('%Y-%m-%d')
                    start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
                    
                logger.info(f"Kullanılan tarih aralığı: {start_date} - {end_date}")
                    
            except Exception as e:
                logger.error(f"Tarih aralığı parse hatası: {str(e)}")
                # Hata durumunda varsayılan tarih aralığını kullan
                end_date = datetime.now().strftime('%Y-%m-%d')
                start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
                logger.info(f"Hata sonrası varsayılan tarih aralığı: {start_date} - {end_date}")
                
        logger.info(f"SQL'e gönderilen tarihler: {start_date}, {end_date}")
        logger.info(f"Tarih filtresi: {start_date} 00:00:00 ile {end_date} 23:59:59 arasında")

        # WHERE koşulunu oluştur
        where_conditions = []
        params = []

        if user_id:
            where_conditions.append(User.user_id == user_id)
        if department:
            # Department ile join yapmak gerekiyor, bu yüzden ayrı bir kontrol yapacağız
            pass  # Bu sorguyla ilgili filtre rapor fonksiyonlarında handle edilecek
        if occupation:
            where_conditions.append(User.occupation == occupation)
        if egitim_durumu:
            where_conditions.append(User.education_level == egitim_durumu)
        if work_type:
            where_conditions.append(User.work_type == work_type)
        if employment_status:
            where_conditions.append(User.employment_status == employment_status)

        # Department filtresini ayrı parametre olarak gönder
        department_filter = department if department else None

        # Rapor verilerini al
        data = None
        title = ""
        
        try:
            if report_type == 'personnel_list':
                data = generate_personnel_list_report(where_conditions, params, department_filter)
                title = "Personel Listesi Raporu"
            elif report_type == 'detailed_attendance':
                data = generate_detailed_attendance_report(where_conditions, params, start_date, end_date, department_filter)
                title = "Dönemsel Ayrıntılı Giriş-Çıkış Raporu"
            elif report_type == 'detailed_timesheet':
                data = generate_detailed_timesheet_report(where_conditions, params, start_date, end_date, department_filter)
                title = "Dönemsel Ayrıntılı Puantaj Raporu"
            elif report_type == 'weekly_timesheet':
                data = generate_weekly_timesheet_report(where_conditions, params, start_date, end_date, department_filter)
                title = "Ayrıntılı Haftalık Puantaj Raporu"
            else:  # daily_timesheet
                data = generate_daily_timesheet_report(where_conditions, params, start_date, end_date, department_filter)
                title = "Günlük Puantaj Raporu"
        except Exception as e:
            print(f"Rapor oluşturma hatası ({report_type}): {str(e)}")
            raise

        # Veriyi JSON için hazırla
        formatted_data = []
        if data:
            # Her satırı dict olarak işle
            for row in data:
                # DictRow nesnesini normal dict'e dönüştür
                if hasattr(row, '_asdict'):
                    row_dict = row._asdict()
                elif isinstance(row, dict):
                    row_dict = row
                else:
                    # Liste veya tuple ise, sıralı olarak işle
                    row_dict = {}
                    for idx, value in enumerate(row):
                        row_dict[f"column_{idx}"] = value
                
                # Özel tipleri (timedelta vb.) dönüştür ve Türkçe başlıkları kullan
                processed_row = {}
                for key, value in row_dict.items():
                    # Değeri dönüştür
                    processed_value = convert_timedelta(value)
                    
                    # NULL veya boş değerleri kontrol et
                    if processed_value is None or processed_value == '' or processed_value == 'None':
                        processed_value = '-'
                        
                    # Türkçe başlığı bul veya orijinal anahtarı kullan
                    header_key = COLUMN_HEADERS.get(key, key.replace('_', ' ').title())
                    processed_row[header_key] = processed_value
                
                formatted_data.append(processed_row)

        # Veri boş mu kontrol et
        if not formatted_data:
            return jsonify({
                'success': False,
                'message': 'Seçilen kriterlere uygun rapor verisi bulunamadı.'
            }), 404

        # HTML tablosunu oluştur
        html = render_template('reports/report_table.html', 
                            data=formatted_data, 
                            title=title,
                            punch_types={'in': 'Giriş', 'out': 'Çıkış'},
                            status_types={0: 'Normal', 1: 'Normal', 4: 'Kart', 5: 'Manuel Giriş', 6: 'Manuel Çıkış'})

        return jsonify({
            'success': True,
            'html': html,
            'title': title,
            'data': formatted_data
        })

    except Exception as e:
        print(f"Genel hata: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f"Rapor oluşturulurken bir hata oluştu: {str(e)}"
        }), 500

    finally:
        db_session.remove()

@reports.route('/download', methods=['POST'])
@login_required
@admin_required
def download():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Veri bulunamadı!'}), 400
            
        file_type = data.get('file_type')
        report_type = data.get('report_type')
        report_data = data.get('report_data')
        
        if not all([file_type, report_type, report_data]):
            return jsonify({'error': 'Eksik parametreler!'}), 400
            
        # Tarih aralığı varsa dosya adına ekle
        date_suffix = ''
        if 'start_date' in data and 'end_date' in data and data['start_date'] and data['end_date']:
            try:
                start = datetime.strptime(data['start_date'], '%Y-%m-%d')
                end = datetime.strptime(data['end_date'], '%Y-%m-%d')
                date_suffix = f"_{start.strftime('%d-%m')}_{end.strftime('%d-%m')}"
            except ValueError:
                # Tarih formatı hatalıysa suffix ekleme
                logger.warning("Tarih formatı hatalı, dosya adına tarih eklenmedi")
                date_suffix = ''
            
        # Timestamp oluştur
        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M')
        
        # Dosya adını oluştur
        base_filename = f"{REPORT_TITLES.get(report_type, 'Rapor')}{date_suffix}_{timestamp}"
        
        if file_type == 'excel':
            output = generate_excel(report_data, report_type)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = f"{base_filename}.xlsx"
        elif file_type == 'pdf':
            output = generate_pdf(report_data, report_type)
            mimetype = 'application/pdf'
            filename = f"{base_filename}.pdf"
        else:
            return jsonify({'error': 'Geçersiz dosya tipi!'}), 400
            
        # Dosyayı gönder
        response = make_response(output)
        response.headers['Content-Type'] = mimetype
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        print(f"Rapor indirme hatası: {str(e)}")
        return jsonify({'error': f'Rapor oluşturulurken bir hata oluştu: {str(e)}'}), 500

def get_group_numbers():
    """Returns a list of group numbers from 0 to 10"""
    return list(range(11))  # 0'dan 10'a kadar

@reports.route('/get_groups', methods=['GET'])
@login_required
@admin_required
def get_groups():
    """Returns available group numbers"""
    groups = get_group_numbers()
    return jsonify({'groups': groups})

def generate_personnel_list_report(where_conditions, params, department_filter):
    try:
        query = db_session.query(User)\
            .filter(User.deleted_at.is_(None))

        # Where koşullarını ekle
        if where_conditions:
            for condition in where_conditions:
                query = query.filter(condition)
                
        # Department filtresi ekle
        if department_filter:
            query = query.outerjoin(Department, User.department_id == Department.department_id)\
                         .filter(Department.department_name == department_filter)

        # Sıralama
        query = query.order_by(User.name)

        users = query.all()

        # Verileri düzenle
        formatted_data = []
        for user in users:
            formatted_row = {
                'Sicil No': user.user_id,
                'Ad Soyad': user.name,
                'Kart No': user.card_no or '-',
                'Departman': user.department.department_name if user.department else '-',
                'Meslek': user.occupation or '-',
                'İş Tanımı': user.job_title or '-',
                'Çalışma Tipi': user.work_type or '-',
                'Çalışma Durumu': user.employment_status or '-'
            }
            formatted_data.append(formatted_row)
            
        return formatted_data

    except Exception as e:
        logger.error(f"Personel listesi raporu hatası: {str(e)}")
        raise

def generate_detailed_attendance_report(where_conditions, params, start_date, end_date, department_filter):
    try:
        # Basit ve düzgün çalışan ana sorgu
        query = db_session.query(
            User.user_id.label('user_id'),
            User.name.label('personel_adi'),
            func.to_char(Record.timestamp, 'DD.MM.YYYY').label('tarih'),
            func.to_char(Record.timestamp, 'HH24:MI').label('saat'),
            case(
                [
                    (Record.punch == 0, 'GİRİŞ'),
                    (Record.punch == 1, 'ÇIKIŞ'),
                    (cast(Record.punch, String) == '0', 'GİRİŞ'),
                    (cast(Record.punch, String) == '1', 'ÇIKIŞ'),
                    (Record.status == 'Manuel Giriş', 'GİRİŞ (Manuel)'),
                    (Record.status == 'Manuel Çıkış', 'ÇIKIŞ (Manuel)'),
                    (Record.status == 'Geç Giriş', 'GİRİŞ (Geç)'),
                    (Record.status == 'Eksik Hareket', 'ÇIKIŞ (Eksik)')
                ],
                else_=func.coalesce(Record.status, 'Bilinmiyor')
            ).label('islem'),
            func.coalesce(Device.name, 'Bilinmiyor').label('kayit_tipi'),
            func.coalesce(Record.status, '-').label('Açıklama')  # description yerine status kullanıyoruz
        ).select_from(Record).join(
            User,
            Record.user_id == User.user_id
        ).outerjoin(
            Device,
            Record.device_id == Device.device_id
        )
        
        # Temel filtreleri ekle - Tarih aralığını tam olarak dahil et
        filters = [
            Record.timestamp >= start_date + ' 00:00:00',
            Record.timestamp <= end_date + ' 23:59:59',
            User.deleted_at.is_(None),
            Record.deleted_at.is_(None)
        ]
        
        # Where koşullarını ekle
        if where_conditions:
            filters.extend(where_conditions)
            
        # Department filtresi ekle
        if department_filter:
            query = query.outerjoin(Department, User.department_id == Department.department_id)
            filters.append(Department.department_name == department_filter)
            
        query = query.filter(and_(*filters)).order_by(Record.timestamp.desc())

        results = query.all()
        
        if not results:
            logger.info("Detaylı giriş-çıkış raporu: Veri bulunamadı")
            return []
            
        # Sonuçları düzenle
        formatted_results = []
        for result in results:
            formatted_row = {}
            
            # Her bir alanı güvenli şekilde işle
            formatted_row[COLUMN_HEADERS.get('user_id', 'Sicil No')] = str(result.user_id) if result.user_id else '-'
            formatted_row[COLUMN_HEADERS.get('personel_adi', 'Ad Soyad')] = str(result.personel_adi) if result.personel_adi else '-'
            formatted_row[COLUMN_HEADERS.get('tarih', 'Tarih')] = str(result.tarih) if result.tarih else '-'
            formatted_row[COLUMN_HEADERS.get('saat', 'Saat')] = str(result.saat) if result.saat else '-'
            formatted_row[COLUMN_HEADERS.get('islem', 'İşlem')] = str(result.islem) if result.islem else '-'
            formatted_row[COLUMN_HEADERS.get('kayit_tipi', 'Terminal')] = str(result.kayit_tipi) if result.kayit_tipi else 'Bilinmiyor'
            formatted_row[COLUMN_HEADERS.get('Açıklama', 'Açıklama')] = str(result.Açıklama) if result.Açıklama and result.Açıklama != 'None' else '-'
            
            formatted_results.append(formatted_row)

        logger.info(f"Detaylı giriş-çıkış raporu: {len(formatted_results)} kayıt bulundu")
        return formatted_results

    except Exception as e:
        logger.error(f"Detaylı giriş-çıkış raporu hatası: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

def generate_detailed_timesheet_report(where_conditions, params, start_date, end_date, department_filter):
    try:
        # Alt sorgu: Günlük kayıtlar - sadece mevcut alanları kullanıyoruz
        daily_records = db_session.query(
            Record.user_id,
            func.date(Record.timestamp).label('work_date'),
            func.min(case(
                [(Record.punch == 0, Record.timestamp)],
                else_=None
            )).label('first_in'),
            func.max(case(
                [(Record.punch == 1, Record.timestamp)],
                else_=None
            )).label('last_out'),
            func.count(Record.id).label('entry_count')
        ).filter(
            Record.timestamp >= start_date + ' 00:00:00',
            Record.timestamp <= end_date + ' 23:59:59',
            Record.deleted_at.is_(None)
        ).group_by(
            Record.user_id,
            func.date(Record.timestamp)
        ).subquery()

        # Ana sorgu
        query = db_session.query(
            User.user_id.label('Sicil No'),
            User.name.label('Ad Soyad'),
            func.coalesce(User.card_no, '-').label('Kart No'),
            func.coalesce(Department.department_name, '-').label('Departman'),
            func.coalesce(User.occupation, '-').label('Meslek'),
            func.coalesce(User.job_title, '-').label('İş Tanımı'),
            func.coalesce(User.work_type, '-').label('Çalışma Tipi'),
            func.coalesce(User.employment_status, '-').label('Çalışma Durumu'),
            func.to_char(daily_records.c.work_date, 'DD.MM.YYYY').label('Tarih'),
            func.to_char(daily_records.c.first_in, 'HH24:MI').label('İlk Giriş'),
            func.to_char(daily_records.c.last_out, 'HH24:MI').label('Son Çıkış'),
            func.extract('epoch', daily_records.c.last_out - daily_records.c.first_in).label('Toplam Süre'),
            daily_records.c.entry_count.label('Kayıt Sayısı')
        ).join(
            daily_records,
            User.user_id == daily_records.c.user_id
        ).outerjoin(
            Department,
            User.department_id == Department.department_id
        ).filter(
            User.deleted_at.is_(None)
        )
        
        # Filtreleri oluştur
        filters = []
        if where_conditions:
            filters.extend(where_conditions)
            
        # Department filtresi ekle
        if department_filter:
            filters.append(Department.department_name == department_filter)
            
        if filters:
            query = query.filter(and_(*filters))
            
        query = query.order_by(
            daily_records.c.work_date.desc(),
            User.name
        )

        results = query.all()

        # Sonuçları düzenle
        formatted_results = []
        for result in results:
            row_dict = result._asdict()
            formatted_row = {}
            for key, value in row_dict.items():
                if key == 'Toplam Süre' and value:
                    hours = int(value) // 3600
                    minutes = (int(value) % 3600) // 60
                    formatted_row[key] = f"{hours:02d}:{minutes:02d}"
                else:
                    formatted_row[key] = str(value) if value is not None else '-'
            formatted_results.append(formatted_row)

        return formatted_results

    except Exception as e:
        logger.error(f"Detaylı puantaj raporu hatası: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

def generate_weekly_timesheet_report(where_conditions, params, start_date, end_date, department_filter):
    try:
        # Alt sorgu: Haftalık kayıtlar - basitleştirilmiş
        weekly_records = db_session.query(
            Record.user_id,
            func.date(Record.timestamp).label('work_date'),
            func.min(case(
                [(Record.punch == 0, Record.timestamp)],
                else_=None
            )).label('first_in'),
            func.max(case(
                [(Record.punch == 1, Record.timestamp)],
                else_=None
            )).label('last_out'),
            func.count(Record.id).label('entry_count'),
            extract('week', Record.timestamp).label('week_number'),
            func.date_trunc('week', Record.timestamp).label('week_start')
        ).filter(
            Record.timestamp >= start_date + ' 00:00:00',
            Record.timestamp <= end_date + ' 23:59:59',
            Record.deleted_at.is_(None)
        ).group_by(
            Record.user_id,
            func.date(Record.timestamp),
            extract('week', Record.timestamp),
            func.date_trunc('week', Record.timestamp)
        ).subquery()

        # Ana sorgu
        query = db_session.query(
            User.user_id.label('Sicil No'),
            User.name.label('Ad Soyad'),
            func.coalesce(User.card_no, '-').label('Kart No'),
            func.coalesce(Department.department_name, '-').label('Departman'),
            func.coalesce(User.occupation, '-').label('Meslek'),
            func.coalesce(User.job_title, '-').label('İş Tanımı'),
            func.coalesce(User.work_type, '-').label('Çalışma Tipi'),
            func.coalesce(User.employment_status, '-').label('Çalışma Durumu'),
            weekly_records.c.week_number.label('Hafta No'),
            func.min(weekly_records.c.week_start).label('Hafta Başlangıç'),
            func.min(weekly_records.c.first_in).label('İlk Giriş'),
            func.max(weekly_records.c.last_out).label('Son Çıkış'),
            func.sum(
                func.extract('epoch', weekly_records.c.last_out - weekly_records.c.first_in)
            ).label('Toplam Süre'),
            func.count(func.distinct(weekly_records.c.work_date)).label('Çalışılan Gün'),
            func.sum(weekly_records.c.entry_count).label('Kayıt Sayısı')
        ).join(
            weekly_records,
            User.user_id == weekly_records.c.user_id
        ).outerjoin(
            Department,
            User.department_id == Department.department_id
        ).filter(
            User.deleted_at.is_(None)
        )
        
        # Filtreleri oluştur
        filters = []
        if where_conditions:
            filters.extend(where_conditions)
            
        # Department filtresi ekle
        if department_filter:
            filters.append(Department.department_name == department_filter)
            
        if filters:
            query = query.filter(and_(*filters))
            
        query = query.group_by(
            User.user_id,
            User.name,
            User.card_no,
            Department.department_name,
            User.occupation,
            User.job_title,
            User.work_type,
            User.employment_status,
            weekly_records.c.week_number
        ).order_by(
            weekly_records.c.week_number.desc(),
            User.name
        )

        results = query.all()

        # Sonuçları düzenle
        formatted_results = []
        for result in results:
            row_dict = result._asdict()
            formatted_row = {}
            for key, value in row_dict.items():
                if key == 'Toplam Süre' and value:
                    hours = int(value) // 3600
                    minutes = (int(value) % 3600) // 60
                    formatted_row[key] = f"{hours:02d}:{minutes:02d}"
                else:
                    formatted_row[key] = str(value) if value is not None else '-'
            formatted_results.append(formatted_row)

        return formatted_results

    except Exception as e:
        logger.error(f"Haftalık puantaj raporu hatası: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

def generate_daily_timesheet_report(where_conditions, params, start_date, end_date, department_filter):
    try:
        logger.info(f"Günlük puantaj raporu başlatılıyor: {start_date} - {end_date}")
        
        # Günlük kayıtlar alt sorgusu - daha güvenli versiyon
        daily_records = db_session.query(
            Record.user_id,
            func.date(Record.timestamp).label('work_date'),
            func.min(
                case(
                    [(and_(Record.punch.in_([0, '0']), Record.timestamp.is_not(None)), Record.timestamp)],
                    else_=None
                )
            ).label('first_in'),
            func.max(
                case(
                    [(and_(Record.punch.in_([1, '1']), Record.timestamp.is_not(None)), Record.timestamp)],
                    else_=None
                )
            ).label('last_out'),
            func.count(Record.id).label('entry_count')
        ).filter(
            and_(
                Record.timestamp >= start_date + ' 00:00:00',
                Record.timestamp <= end_date + ' 23:59:59',
                Record.deleted_at.is_(None),
                Record.timestamp.is_not(None)
            )
        ).group_by(
            Record.user_id,
            func.date(Record.timestamp)
        ).subquery()

        # Ana sorgu - temiz ve basit
        query = db_session.query(
            User.user_id.label('Sicil No'),
            User.name.label('Ad Soyad'),
            func.coalesce(User.card_no, '-').label('Kart No'),
            func.coalesce(Department.department_name, '-').label('Departman'),
            func.coalesce(User.occupation, '-').label('Meslek'),
            func.coalesce(User.job_title, '-').label('İş Tanımı'),
            func.coalesce(User.work_type, '-').label('Çalışma Tipi'),
            func.coalesce(User.employment_status, '-').label('Çalışma Durumu'),
            func.to_char(daily_records.c.work_date, 'DD.MM.YYYY').label('Tarih'),
            func.coalesce(
                func.to_char(daily_records.c.first_in, 'HH24:MI'),
                '-'
            ).label('İlk Giriş'),
            func.coalesce(
                func.to_char(daily_records.c.last_out, 'HH24:MI'),
                '-'
            ).label('Son Çıkış'),
            func.coalesce(daily_records.c.entry_count, 0).label('Kayıt Sayısı'),
            # Toplam süre hesaplama
            case(
                [
                    (
                        and_(
                            daily_records.c.first_in.is_not(None),
                            daily_records.c.last_out.is_not(None)
                        ),
                        func.extract('epoch', daily_records.c.last_out - daily_records.c.first_in)
                    )
                ],
                else_=0
            ).label('Toplam Süre'),
            # Durum hesaplama
            case(
                [
                    (
                        and_(
                            daily_records.c.first_in.is_not(None),
                            daily_records.c.last_out.is_not(None),
                            func.extract('epoch', daily_records.c.last_out - daily_records.c.first_in) >= 14400
                        ),
                        'Tam Gün'
                    ),
                    (
                        and_(
                            daily_records.c.first_in.is_not(None),
                            daily_records.c.last_out.is_not(None),
                            func.extract('epoch', daily_records.c.last_out - daily_records.c.first_in) < 14400
                        ),
                        'Yarım Gün'
                    ),
                    (
                        or_(
                            daily_records.c.first_in.is_(None),
                            daily_records.c.last_out.is_(None)
                        ),
                        'Eksik Kayıt'
                    )
                ],
                else_='Belirsiz'
            ).label('Durum'),
            # Basit mesai durumu
            case(
                [
                    (
                        and_(
                            daily_records.c.first_in.is_not(None),
                            daily_records.c.last_out.is_not(None)
                        ),
                        'Normal'
                    )
                ],
                else_='Eksik'
            ).label('Mesai Durumu')
        ).select_from(
            User
        ).join(
            daily_records,
            User.user_id == daily_records.c.user_id
        ).outerjoin(
            Department,
            User.department_id == Department.department_id
        ).filter(
            User.deleted_at.is_(None)
        )
        
        # Filtreleri ekle
        if where_conditions:
            query = query.filter(and_(*where_conditions))
            
        # Department filtresi
        if department_filter:
            query = query.filter(Department.department_name == department_filter)
            
        # Sıralama
        query = query.order_by(
            daily_records.c.work_date.desc(),
            User.name.asc()
        )

        logger.info("SQL sorgusu hazırlandı, çalıştırılıyor...")
        results = query.all()
        
        if not results:
            logger.info("Günlük puantaj raporu: Veri bulunamadı")
            return []

        logger.info(f"Günlük puantaj raporu: {len(results)} kayıt bulundu")

        # Sonuçları güvenli şekilde işle
        formatted_results = []
        for result in results:
            try:
                row_dict = result._asdict()
                formatted_row = {}
                
                for key, value in row_dict.items():
                    if key == 'Toplam Süre' and value and value != 0:
                        # Saniye cinsinden değeri saat:dakika formatına çevir
                        total_seconds = int(float(value))
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        formatted_row[key] = f"{hours:02d}:{minutes:02d}"
                    elif value is None:
                        formatted_row[key] = '-'
                    elif str(value).lower() in ('none', 'null'):
                        formatted_row[key] = '-'
                    else:
                        formatted_row[key] = str(value)
                        
                formatted_results.append(formatted_row)
                
            except Exception as row_error:
                logger.error(f"Satır işleme hatası: {str(row_error)}")
                # Hatalı satırı atla, devam et
                continue

        logger.info(f"Günlük puantaj raporu başarıyla tamamlandı: {len(formatted_results)} kayıt")
        return formatted_results

    except Exception as e:
        logger.error(f"Günlük puantaj raporu hatası: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

def generate_excel(data, report_type):
    try:
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        
        # Stil tanımlamaları
        header_font = Font(name='Arial', bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlıkları yaz
        headers = list(data[0].keys())
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = border
        
        # Verileri yaz
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, (key, value) in enumerate(row_data.items(), start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # HTML içeriği temizle
                if key == "Açıklama" and isinstance(value, str):
                    cell.value = strip_html_tags(value)
                else:
                    cell.value = str(value)
                cell.alignment = cell_alignment
                cell.border = border
        
        # Sütun genişliklerini ayarla
        for col in ws.columns:
            max_length = 0
            column = [cell for cell in col]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        # Excel dosyasını kaydet
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    except Exception as e:
        print(f"Excel oluşturma hatası: {str(e)}")
        raise

def generate_pdf(data, report_type):
    try:
        # PDF buffer oluştur
        buffer = BytesIO()
        
        # PDF dokümanı oluştur
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        # Stil tanımlamaları
        styles = getSampleStyleSheet()
        
        # Arial Unicode MS fontunu kullan (Windows'ta varsayılan olarak yüklü)
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName='Helvetica',  # Helvetica fontu Türkçe karakterleri destekler
            fontSize=10,
            leading=12
        )
        
        # Başlık stili
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=14,
            leading=16,
            alignment=1  # Ortalı
        )
        
        # Rapor başlığını belirle
        title_mapping = {
            'personnel_list': 'Personel Listesi',
            'detailed_attendance': 'Detaylı Giriş-Çıkış Raporu',
            'detailed_timesheet': 'Detaylı Puantaj Raporu',
            'weekly_timesheet': 'Haftalık Puantaj Raporu',
            'daily_timesheet': 'Günlük Puantaj Raporu'
        }
        
        # Rapor başlığını oluştur
        title = title_mapping.get(report_type, 'Rapor')
        elements = [Paragraph(title, title_style)]
        elements.append(Spacer(1, 20))
        
        if not data:
            elements.append(Paragraph('Rapor verisi bulunamadı.', normal_style))
        else:
            # Tablo verilerini hazırla
            table_data = []
            
            # Başlık satırını ekle
            headers = [Paragraph(str(key).replace('_', ' ').title(), normal_style) for key in data[0].keys()]
            table_data.append(headers)
            
            # Verileri ekle
            for row in data:
                table_row = []
                for key, value in row.items():
                    # None değerleri kontrol et
                    if value is None:
                        cell_value = '-'
                    else:
                        # HTML etiketlerini temizle (Açıklama sütunu için)
                        if key == "Açıklama" and isinstance(value, str):
                            cell_value = strip_html_tags(value)
                        else:
                            cell_value = str(value)
                    table_row.append(Paragraph(cell_value, normal_style))
                table_data.append(table_row)
            
            # Tablo stilini oluştur
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ])
            
            # Tabloyu oluştur
            table = Table(table_data)
            table.setStyle(table_style)
            
            # Tabloyu ekle
            elements.append(table)
        
        # PDF oluştur
        doc.build(elements)
        
        # Buffer'ı başa al
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        print(f"PDF oluşturma hatası: {str(e)}")
        raise 

@reports.route('/api/active-users-count', methods=['GET'])
@login_required
@admin_required
def active_users_count():
    try:
        today = datetime.now().date()
        
        # Son hareketleri al ve içeride olanları say
        subquery = db_session.query(
            Record.user_id,
            Record.timestamp,
            Record.punch,
            Record.status,
            func.row_number().over(
                partition_by=Record.user_id,
                order_by=Record.timestamp.desc()
            ).label('rn')
        ).filter(
            func.date(Record.timestamp) == today
        ).subquery()

        active_count = db_session.query(func.count(subquery.c.user_id))\
            .filter(
                subquery.c.rn == 1,
                or_(
                    subquery.c.punch == '1',
                    cast(subquery.c.punch, String) == '1',
                    subquery.c.status.in_(['5', 'Manuel Giriş'])
                )
            ).scalar()
        
        return jsonify({
            'success': True, 
            'active_count': active_count,
            'timestamp': datetime.now().strftime('%H:%M:%S')
        })
        
    except Exception as e:
        logger.error(f"Aktif kullanıcı sayısı hesaplama hatası: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        db_session.remove() 